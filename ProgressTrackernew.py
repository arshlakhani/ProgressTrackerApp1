import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import sqlite3
import time
from datetime import datetime
from openpyxl import *

TrackerPath = 'trackingnew.xlsx'
ProgressTrackerPath = 'progress_tracker.db'
TrackerDf = pd.read_excel(TrackerPath)

steps = [
    {"name": "Application submission in o/o DTCP", "approval": ["RECORD"],
     "sub_steps": ["Application submission in o/o DTCP"]},
    {"name": "Scrutiny of Documents by o/o DTCP", "approval": ["JD", "PA", "ATP", "DTP", "STP"],
     "sub_steps": ["Scrutiny of Documents by o/o DTCP"]},
    {
        "name": "Letter forwarding circulation of BPs to DTP & STP(Circle), SE-HSVP, Fire officer-ULB, PKL and Observations letter issued",
        "sub_steps": [
            "Letter forwarding circulation of BPs to DTP & STP(Circle), SE-HSVP, Fire officer-ULB, PKL and Observations letter issued"],
        "approval": ["STP(HQL)"]},
    {"name": "Examination by Concerned Circle - District Town Planner office",
     "sub_steps": ["Examination by Concerned Circle - District Town Planner office", "Compilation of observations"],
     "approval": ["JD", "SD", "PA", "ATP", "DTP"]},
    {"name": "Examination by Concerned Circle - Senior Town Planner office",
     "sub_steps": ["Examination by Concerned Circle - Senior Town Planner office", "Compilation of observations",
                   "Site reports receival"],
     "approval": ["JD", "ATP", "STP"]},
    {"name": "Examination by SE / Executive Engineer - HSVP",
     "sub_steps": ["Examination by SE / Executive Engineer - HSVP", "Compilation of observations",
                   "Site reports receival"],
     "approval": ["JD", "SDM", "HDM", "CHD", "SDE", "CHD", "SDO", "XEN", "SE", "CE", "SE"]},
    {"name": "Examination by Fire Officer, Urban Local Bodies",
     "sub_steps": ["Examination by Fire Officer, Urban Local Bodies", "Compilation of observations",
                   "Site reports receival"],
     "approval": ["ASST.", "SUPDT.", "CONSULTANT", "FIRE OFFICER"]},
    {"name": "Compilation of Reports in o/o DTCP 1", "sub_steps": ["Compilation of Reports in o/o DTCP"],
     "approval": ["JD", "PA", "ATP", "DTP", "ARCHITECT", "STP", "ARCHITECT", "CTP", "DTP"]},
    {"name": "Fixing of Meeting of BPAC to review the comments / report of all above",
     "sub_steps": ["Fixing of Meeting of BPAC to review the comments / report of all above"], "approval": ["STP"]},
    {"name": "Plan reviewed in BPAC Committee", "sub_steps": ["Plan reviewed in BPAC Committee"],
     "approval": ["o/o DTCP"]},
    {"name": "Observations conveyed", "sub_steps": ["Observations conveyed"],
     "approval": [""]},
    {"name": "Resubmission of Dwgs after compliance", "sub_steps": ["Resubmission of Dwgs after compliance"],
     "approval": [""]},
    {"name": "Circulation of BPs to STP (circle)GGN, SE-HSVP, Fire officer-ULB, Pkl for signatures.",
     "sub_steps": ["Circulation of BPs to STP (circle)GGN, SE-HSVP, Fire officer-ULB, Pkl for signatures."],
     "approval": [""]},
    {"name": "Examination of BPs at Field offices", "sub_steps": ["Examination of BPs at Field offices"],
     "approval": [""]},
    {"name": "Compilation of Reports in o/o DTCP", "sub_steps": ["Compilation of Reports in o/o DTCP"],
     "approval": ["JD"]},
    {"name": "Verification of the Licence / CLU permission / pending dues by the Department",
     "sub_steps": ["Verification of the Licence / CLU permission / pending dues by the Department", "sub_steps"],
     "approval": ["SO", "AO", "CAO"]},
    {"name": "Approved Building Plans & BR-III issued", "sub_steps": ["Approved Building Plans & BR-III issued"],
     "approval": ["JD", "ATP", "DTP", "ARCHITECT", "STP", "CTP"]}
]
conn = sqlite3.connect(ProgressTrackerPath)
cursor = conn.cursor()

cursor.execute('''
CREATE TABLE IF NOT EXISTS projectList (
    id INTEGER PRIMARY KEY,
    name TEXT,
    desc TEXT,
    created_time TEXT,
    estimated_time TEXT
)
''')
conn.commit()

root = tk.Tk()
root.title("Projects")
root.geometry('600x600')


def showProjectList():
    cursor.execute("SELECT name FROM projectList")
    rows = cursor.fetchall()
    project_names = [row[0] for row in rows]

    window = tk.Toplevel(root)
    window.title("Project List")

    for project_name in project_names:
        newButton = tk.Button(window, text=project_name, command=lambda name=project_name: openProject(name))
        newButton.pack()


def uploadDocument(approval_name, step_name, project_name):
    file_path = filedialog.askopenfilename()
    if file_path:
        cursor.execute(f'''
        UPDATE progress_{project_name} 
        SET document_path = ?
        WHERE step = ? AND sub_step = ?
        ''', (file_path, step_name, approval_name))
        conn.commit()
        messagebox.showinfo("Info", "Document uploaded successfully!")


def openStepDetails(project_name, step_name, sub_steps, approvals):
    def markCompleted(step, approval, completed, submitted_date, completed_date):
        cursor.execute(f'''
        UPDATE progress_{project_name} 
        SET completed = ?, submitted_date = ?, completed_date = ?
        WHERE step = ? AND sub_step = ?
        ''', (completed, submitted_date, completed_date, step, approval))
        conn.commit()

    def saveProgress():
        for checkbox, approval, submitted_entry, completed_entry in approval_vars:
            submitted_date = submitted_entry.get() if submitted_entry.get() else None
            completed_date = completed_entry.get() if completed_entry.get() else None
            markCompleted(step_name, approval, checkbox.get(), submitted_date, completed_date)

        messagebox.showinfo("Info", "Progress Saved")

    window = tk.Toplevel(root)
    window.title(f"Step Details: {step_name}")

    step_frame = tk.LabelFrame(window, text="Sub-Steps")
    step_frame.pack(fill="both")

    for sub_step in sub_steps:
        var = tk.BooleanVar()
        checkbox = tk.Checkbutton(step_frame, text=sub_step, variable=var)
        checkbox.pack()

    approval_frame = tk.LabelFrame(window, text="Approvals")
    approval_frame.pack(fill="both")

    approval_vars = []
    for approval in approvals:
        var = tk.BooleanVar()
        submitted_entry = tk.Entry(approval_frame)
        completed_entry = tk.Entry(approval_frame)

        cursor.execute(f'''
        SELECT completed, submitted_date, completed_date, document_path FROM progress_{project_name} 
        WHERE step = ? AND sub_step = ?
        ''', (step_name, approval))
        result = cursor.fetchone()
        if result:
            var.set(result[0])
            if result[1]:
                submitted_entry.insert(0, result[1])
            if result[2]:
                completed_entry.insert(0, result[2])
            document_path = result[3]
        else:
            document_path = None
            cursor.execute(f'''
            INSERT INTO progress_{project_name} (step, sub_step, created_time, target_time, actual_time, document_path, completed, completed_time, submitted_date, completed_date) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (step_name, approval, time.time(), time.time() + 7.862e+6, None, None, False, None, None, None))
            conn.commit()

        checkbox = tk.Checkbutton(approval_frame, text=approval, variable=var)
        checkbox.pack()
        submitted_entry.pack()
        completed_entry.pack()

        upload_button = tk.Button(approval_frame, text="Upload Document",
                                  command=lambda a=approval: uploadDocument(a, step_name, project_name))
        upload_button.pack()

        if document_path:
            doc_link = tk.Label(approval_frame, text="Document", fg="blue", cursor="hand2")
            doc_link.pack()
            doc_link.bind("<Button-1>", lambda e, path=document_path: open_document(path))

        approval_vars.append((var, approval, submitted_entry, completed_entry))

    save_button = tk.Button(window, text="Save Progress", command=saveProgress)
    save_button.pack()


def open_document(file_path):
    import os
    os.startfile(file_path)


def openProject(project_name):
    window = tk.Toplevel(root)
    window.title(f"Project: {project_name}")

    for step in steps:
        step_button = tk.Button(window, text=step['name'],
                                command=lambda step=step: openStepDetails(project_name, step['name'], step['sub_steps'],
                                                                          step['approval']))
        step_button.pack(fill="both")

    download_button = tk.Button(window, text="Download Progress", command=lambda: downloadProgress(project_name))
    download_button.pack(fill="both")


def projectTop():
    def addProject(name, desc, topl):
        cursor.execute(f"INSERT INTO projectList (name, desc, created_time, estimated_time) VALUES (?, ?, ?, ?)",
                       (name, desc, time.time(), time.time() + 7.862e+6))
        conn.commit()
        project_name = str(name).replace(" ", "_")
        cursor.execute(f'''
        CREATE TABLE IF NOT EXISTS progress_{project_name} (
            id INTEGER PRIMARY KEY,
            step TEXT,
            sub_step TEXT,
            created_time TEXT,
            target_time TEXT,
            actual_time TEXT,
            document_path TEXT,
            completed BOOLEAN,
            completed_time TEXT,
            submitted_date TEXT,
            completed_date TEXT
        )
        ''')
        conn.commit()
        messagebox.showinfo("Info", f"Project '{name}' added successfully!")

    project_name = tk.StringVar()
    project_desc = tk.StringVar()

    top = tk.Toplevel(root)
    top.title("Add Project")

    tk.Label(top, text="Project Name").pack()
    tk.Entry(top, textvariable=project_name).pack()

    tk.Label(top, text="Description").pack()
    tk.Entry(top, textvariable=project_desc).pack()

    tk.Button(top, text="Add Project", command=lambda: addProject(project_name.get(), project_desc.get(), top)).pack()


def downloadProgress(project_name):
    def swap_rows(ws, row1, row2):
        """
        Swap the values of two rows in a worksheet.
        """
        for col in range(1, ws.max_column + 1):
            cell1 = ws.cell(row=row1, column=col)
            cell2 = ws.cell(row=row2, column=col)
            # Swap values
            cell1.value, cell2.value = cell2.value, cell1.value

    # Create a DataFrame from the steps list
    step_df = pd.DataFrame(steps)

    # Flatten the approvals list into separate rows for each step
    approvals_expanded = []
    for index, row in step_df.iterrows():
        step_name = row['name']
        for approval in row['approval']:
            approvals_expanded.append({'step': step_name, 'approval': approval})

    approvals_df = pd.DataFrame(approvals_expanded)

    # Initialize the data for all steps
    step_df['submitted_date'] = None
    step_df['completed_date'] = None
    step_df['document_path'] = None

    # Fetch data from the database for the given project
    cursor.execute(
        f"SELECT id, step, sub_step, submitted_date, completed_date, document_path FROM progress_{project_name}")
    rows = cursor.fetchall()
    columns = [description[0] for description in cursor.description]
    progress_df = pd.DataFrame(rows, columns=columns)

    # Debug: Print fetched data
    print("Fetched Data:")
    print(progress_df)

    # Merge the steps DataFrame with the fetched data
    merged_df = pd.merge(step_df, progress_df, left_on=['name'], right_on=['step'], how='left',
                         suffixes=('', '_progress'))
    merged_df['submitted_date'] = merged_df['submitted_date_progress'].combine_first(merged_df['submitted_date'])
    merged_df['completed_date'] = merged_df['completed_date_progress'].combine_first(merged_df['completed_date'])
    merged_df['document_path'] = merged_df['document_path_progress'].combine_first(merged_df['document_path'])

    # Select relevant columns
    merged_df = merged_df[['name', 'submitted_date', 'completed_date', 'document_path']]

    # Merge approvals_df with merged_df
    final_df = pd.merge(approvals_df, merged_df, left_on='step', right_on='name', how='left')
    final_df = final_df[['step', 'approval', 'submitted_date', 'completed_date', 'document_path']]

    # Sort the DataFrame based on the order of steps defined in the steps list
    step_order = {step['name']: idx for idx, step in enumerate(steps)}

    # Debug: Print step order mapping
    print("Step Order Mapping:")
    print(step_order)

    # Apply step order mapping
    final_df['step_order'] = final_df['step'].map(step_order)

    # Debug: Print DataFrame with step order
    print("DataFrame with Step Order:")
    print(final_df)

    # Sort the DataFrame by step order
    final_df.sort_values(by='step_order', inplace=True)
    final_df.drop(columns=['step_order'], inplace=True)

    # Debug: Print final DataFrame
    print("Final DataFrame:")
    print(final_df)

    output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if output_path:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            final_df.to_excel(writer, sheet_name='Progress', index=False)

            # Access the openpyxl workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Progress']

            # Define column widths
            worksheet.column_dimensions['A'].width = 40  # Step
            worksheet.column_dimensions['B'].width = 20  # Approval
            worksheet.column_dimensions['C'].width = 20  # Submitted Date
            worksheet.column_dimensions['D'].width = 20  # Completed Date
            worksheet.column_dimensions['E'].width = 20  # Document Path

            stepStart = [2, 3, 8, 9, 13, 14, 16, 17, 18, 28, 32, 33, 34, 35, 36, 37, 38, 58, 61, 67]

            worksheet.insert_cols(1)
            worksheet['A1'] = 'Sr No'
            for i in range(len(stepStart) - 1):
                worksheet[f'A{stepStart[i]}'] = i + 1  # Set the value before merging cells
                # worksheet.merge_cells(f'A{stepStart[i]}:A{stepStart[i + 1] - 1}')
                # worksheet.merge_cells(f'B{stepStart[i]}:B{stepStart[i + 1] - 1}')
            worksheet.insert_cols(4)
            worksheet['A1'] = 'Sr No'
            worksheet['B1'] = 'Approval of BPS'
            worksheet['C1'] = 'Approval Stage'
            worksheet['D1'] = 'Timeline Targeted'
            worksheet['E1'] = 'Submitted Date'
            worksheet['F1'] = 'Completed Date'
            worksheet['D2'] = '2 Weeks'
            worksheet.merge_cells('D2:D8')
            worksheet['D9'] = '3 Weeks'
            worksheet.merge_cells('D9:D34')
            worksheet['D35'] = '1 Week'
            worksheet.merge_cells('D35:D44')
            worksheet['D45'] = '2 Weeks'
            worksheet.merge_cells('D45:D47')
            worksheet['D48'] = '3 Weeks'
            worksheet.merge_cells('D48:D49')
            worksheet['D50'] = '2 Weeks'
            worksheet.merge_cells('D50:D57')

        messagebox.showinfo("Info", "Progress downloaded successfully")


'''
    output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if output_path:
        df.to_excel(output_path, index=False)
        messagebox.showinfo("Info", "Progress downloaded successfully")
        wb = load_workbook(output_path)
        ws = wb.active
        ws = wb["Sheet1"]
        max_row = 29

        for row in range(2, (max_row // 2) + 1):
            swap_rows(ws, row, max_row - row + 1)

        ws.merge_cells('B4:B12')
        ws.merge_cells('B14:B16')
        ws.merge_cells('B17:B22')
        ws.merge_cells('B25:B27')
        ws.merge_cells('A4:A12')
        ws.merge_cells('A14:A16')
        ws.merge_cells('A17:A22')
        ws.merge_cells('A25:A27')
        ws['A13'] = 4
        ws['A14'] = 5
        ws['A17'] = 6
        ws['A23'] = 7
        ws['A24'] = 8
        ws['A25'] = 9
        ws['A28'] = 10
        ws.insert_cols(4)

        '''

add_button = tk.Button(root, text="Add Project", command=projectTop)
add_button.pack()

show_button = tk.Button(root, text="Show Projects", command=showProjectList)
show_button.pack()

root.mainloop()
